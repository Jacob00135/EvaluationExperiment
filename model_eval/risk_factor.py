import os
import sys
import pdb
import json
import openpyxl
import numpy as np
import pandas as pd
import compute_performance as cp
from time import time as get_timestamp
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

now_path = os.path.dirname(__file__)
sys.path.append(os.path.realpath(os.path.join(now_path, '..')))
from config import root_path, category_list


def processing_fhq(fhq, var_name):
    # 处理FHQMOMAD、FHQDADAD字段的函数
    fhq[var_name] = fhq[var_name].map({0: 0, 1: 1, 2: np.nan, -1: np.nan, -4: np.nan})
    var = fhq[['RID', var_name]].dropna(subset=[var_name])
    var.index = range(var.shape[0])
    counter = {}
    for i, rid in enumerate(var['RID'].values):
        if rid not in counter:
            counter[rid] = [i]
        else:
            counter[rid].append(i)
    delete_row_index = []
    for rid, index_list in filter(lambda t: len(t[1]) > 1, counter.items()):
        best_value = 0
        for index in index_list:
            if best_value == 1:
                break
            best_value = var.loc[index, var_name]
        var.loc[index_list[0], var_name] = best_value
        delete_row_index.extend(index_list[1:])
    var = var.drop(delete_row_index, axis=0)
    var.index = range(var.shape[0])
    return var


def preprocess(test_source_path, risk_factor_path, result_save_path):
    """
    将stage、age、genetics、family history、education这些危险因素作为变量加入到测试集中，
    处理后得到的测试集将有以下形状：
    样本量 = 源测试集样本量
    字段有：['stage', 'age', 'genetics', 'family_history', 'education']
    注意：处理过程中使用('RID', 'VISCODE')作为主键进行表合并，可能会出现某个字段值缺失

    :param test_source_path: str。源测试集文件的路径
    :param risk_factor_path: str。包含了风险因素的表所在的目录路径，这些目录必须包含这些文件：
                             ADNIMERGE.csv、FAMXHPAR.csv、FAMXHSIB.csv、FHQ.csv、RECFHQ.csv
    :param result_save_path: str。处理完毕后得到的DataFrame保存的路径，注意必须以.csv为扩展名
    :return: None
    """

    # 载入数据
    test_source = pd.read_csv(test_source_path)
    adnimerge = pd.read_csv(os.path.join(risk_factor_path, 'ADNIMERGE.csv'))
    famxhpar = pd.read_csv(os.path.join(risk_factor_path, 'FAMXHPAR.csv'))
    famxhsib = pd.read_csv(os.path.join(risk_factor_path, 'FAMXHSIB.csv'))
    fhq = pd.read_csv(os.path.join(risk_factor_path, 'FHQ.csv'))
    recfhq = pd.read_csv(os.path.join(risk_factor_path, 'RECFHQ.csv'))

    # 添加stage、age、genetics、education、gender、race、mmse
    data = pd.merge(
        test_source[['RID', 'VISCODE', 'age', 'gender', 'mmse']],
        adnimerge[['RID', 'VISCODE', 'ORIGPROT', 'APOE4', 'PTEDUCAT', 'PTRACCAT']].rename(
            columns={'ORIGPROT': 'stage', 'APOE4': 'genetics', 'PTEDUCAT': 'education', 'PTRACCAT': 'race'}
        ),
        how='left',
        on=['RID', 'VISCODE']
    )
    data['stage'] = data['stage'].map({'ADNI1': 0, 'ADNIGO': 1, 'ADNI2': 2, 'ADNI3': 3}).fillna(-4).values.astype(
        'int32')
    data['age'] = data['age'].fillna(-4).values.astype('int32')
    data['gender'] = data['gender'].map({'female': 0, 'male': 1}).fillna(-4).values.astype('int32')
    data['genetics'] = data['genetics'].fillna(-4).values.astype('int32')
    data['education'] = data['education'].fillna(-4).values.astype('int32')
    data['race'] = data['race'].map({
        'White': 0,
        'Asian': 1,
        'Black': 2,
        'Am Indian/Alaskan': 3,
        'More than one': 4,
        'Hawaiian/Other PI': 5
    }).fillna(-4).astype('int32')
    data['mmse'] = data['mmse'].fillna(-4).astype('int32')

    # 添加family history
    data = pd.merge(data, famxhpar[['RID', 'MOTHAD', 'FATHAD']], how='left', on='RID')
    data = pd.merge(data, famxhsib[['RID', 'SIBAD']], how='left', on='RID')
    data = pd.merge(data, processing_fhq(fhq, 'FHQMOMAD'), how='left', on='RID')
    data = pd.merge(data, processing_fhq(fhq, 'FHQDADAD'), how='left', on='RID')
    data = pd.merge(data, processing_fhq(recfhq, 'FHQSIBAD'), how='left', on='RID')
    data.index = range(data.shape[0])
    fh_source = data[['MOTHAD', 'FATHAD', 'SIBAD', 'FHQMOMAD', 'FHQDADAD', 'FHQSIBAD']].fillna(-4).values.astype(
        'int32')
    family_history = np.zeros(fh_source.shape[0], 'int32')
    for i in range(fh_source.shape[0]):
        if (fh_source[i, :] == 1).sum() > 0:
            family_history[i] = 1
        elif (fh_source[i, :] == 0).sum() > 0:
            family_history[i] = 0
        else:
            family_history[i] = -4
    data['family_history'] = family_history
    data = data.drop(['MOTHAD', 'FATHAD', 'SIBAD', 'FHQMOMAD', 'FHQDADAD', 'FHQSIBAD'], axis=1)
    data.index = range(data.shape[0])

    # 保存
    for c in [*category_list, 'COG']:
        data[c] = test_source[c].values.astype('int32')
    data['benefit'] = test_source['benefit'].fillna(0).values.astype('float32')
    data.to_csv(result_save_path, index=False)


class RiskFactorEval(object):

    def __init__(self, test_set_path, scores_path, result_dir_path):
        self.test_set = pd.read_csv(test_set_path)
        self.scores_dict = {k: np.load(v) for k, v in scores_path.items()}
        self.result_dir_path = result_dir_path

    def grouping_by_stage(self):
        raise ValueError('stage = 3的组样本量仅有4！')
        print('{:=^50}'.format('stage分组'))
        for v in range(4):
            boolean_filter = self.test_set['stage'].values == v
            print('stage = {}\t样本量 = {}'.format(v, boolean_filter.sum()))
            test_set = self.test_set[boolean_filter]
            test_set.index = range(test_set.shape[0])
            for model_name, scores in self.scores_dict.items():
                start_time = get_timestamp()
                eval_result = cp.compute(test_set, scores[:, boolean_filter])
                save_path = os.path.join(self.result_dir_path, 'stage_{}_{}.csv'.format(v, model_name))
                eval_result.to_csv(save_path, index=False)
                print('stage_{}_{}: {:.0f}s'.format(v, model_name, get_timestamp() - start_time))

    def grouping_by_race(self):
        raise ValueError('race = 5的组样本量仅有1！')
        print('{:=^50}'.format('race分组'))
        for v in range(6):
            boolean_filter = self.test_set['race'].values == v
            print('race = {}\t样本量 = {}'.format(v, boolean_filter.sum()))
            test_set = self.test_set[boolean_filter]
            test_set.index = range(test_set.shape[0])
            for model_name, scores in self.scores_dict.items():
                start_time = get_timestamp()
                eval_result = cp.compute(test_set, scores[:, boolean_filter])
                save_path = os.path.join(self.result_dir_path, 'race_{}_{}.csv'.format(v, model_name))
                eval_result.to_csv(save_path, index=False)
                print('race_{}_{}: {:.0f}s'.format(v, model_name, get_timestamp() - start_time))

    def grouping_by_age(self):
        """
        分组
        [min, 65)
        [65, 75)
        [75, 85)
        [85, max]
        """
        print('{:=^50}'.format('age分组'))
        print('  区间    样本量')
        var = self.test_set['age'].values
        age_scale = [65, 75, 85]
        for i in range(len(age_scale) + 1):
            if i == 0:
                min_age = var[var > 0].min()
                max_age = age_scale[0]
                boolean_filter = (0 < var) & (var < max_age)
                print('[{}, {})    {}'.format(min_age, max_age, boolean_filter.sum()))
            elif i == len(age_scale):
                min_age = age_scale[-1]
                max_age = var.max()
                boolean_filter = var >= min_age
                print('[{}, {}]    {}'.format(min_age, max_age, boolean_filter.sum()))
            else:
                min_age = age_scale[i - 1]
                max_age = age_scale[i]
                boolean_filter = (min_age <= var) & (var < max_age)
                print('[{}, {})    {}'.format(min_age, max_age, boolean_filter.sum()))
            test_set = self.test_set[boolean_filter]
            test_set.index = range(test_set.shape[0])
            for model_name, scores in self.scores_dict.items():
                start_time = get_timestamp()
                eval_result = cp.compute(test_set, scores[:, boolean_filter])
                save_filename = 'age_{}_{}_{}.csv'.format(min_age, max_age, model_name)
                eval_result.to_csv(os.path.join(self.result_dir_path, save_filename), index=False)
                print('age_{}_{}_{}: {:.0f}s'.format(min_age, max_age, model_name, get_timestamp() - start_time))

    def grouping_by_gender(self):
        var = self.test_set['gender'].values
        print('{:=^50}'.format('gender分组'))
        for v in range(2):
            boolean_filter = var == v
            print('gender = {}\t样本量 = {}'.format(v, boolean_filter.sum()))
            test_set = self.test_set[boolean_filter]
            test_set.index = range(test_set.shape[0])
            for model_name, scores in self.scores_dict.items():
                start_time = get_timestamp()
                eval_result = cp.compute(test_set, scores[:, boolean_filter])
                save_filename = 'gender_{}_{}.csv'.format(v, model_name)
                eval_result.to_csv(os.path.join(self.result_dir_path, save_filename), index=False)
                print('gender_{}_{}: {:.0f}s'.format(v, model_name, get_timestamp() - start_time))

    def grouping_by_genetics(self):
        var = self.test_set['genetics'].values
        print('{:=^50}'.format('genetics分组'))
        for v in range(3):
            boolean_filter = var == v
            print('apoe4 = {}\t样本量 = {}'.format(v, boolean_filter.sum()))
            test_set = self.test_set[boolean_filter]
            test_set.index = range(test_set.shape[0])
            for model_name, scores in self.scores_dict.items():
                start_time = get_timestamp()
                eval_result = cp.compute(test_set, scores[:, boolean_filter])
                save_filename = 'genetics_{}_{}.csv'.format(v, model_name)
                eval_result.to_csv(os.path.join(self.result_dir_path, save_filename), index=False)
                print('genetics_{}_{}: {:.0f}s'.format(v, model_name, get_timestamp() - start_time))

    def grouping_by_family_history(self):
        var = self.test_set['family_history'].values
        print('{:=^50}'.format('family_history分组'))
        for v in range(2):
            boolean_filter = var == v
            print('family_history = {}\t样本量 = {}'.format(v, boolean_filter.sum()))
            test_set = self.test_set[boolean_filter]
            test_set.index = range(test_set.shape[0])
            for model_name, scores in self.scores_dict.items():
                start_time = get_timestamp()
                eval_result = cp.compute(test_set, scores[:, boolean_filter])
                save_filename = 'family_history_{}_{}.csv'.format(v, model_name)
                eval_result.to_csv(os.path.join(self.result_dir_path, save_filename), index=False)
                print('family_history_{}_{}: {:.0f}s'.format(v, model_name, get_timestamp() - start_time))

    def grouping_by_education(self):
        """
        理想分组
        (-∞, 12)
        [12, 15)
        [15, +∞)
        """
        print('{:=^50}'.format('age分组'))
        print('  区间    样本量')
        var = self.test_set['education'].values
        education_scale = [12, 15]
        for i in range(len(education_scale) + 1):
            if i == 0:
                min_education = var[var > 0].min()
                max_education = education_scale[0]
                boolean_filter = (0 < var) & (var < max_education)
                print('[{}, {})    {}'.format(min_education, max_education, boolean_filter.sum()))
            elif i == len(education_scale):
                min_education = education_scale[-1]
                max_education = var.max()
                boolean_filter = var >= min_education
                print('[{}, {}]    {}'.format(min_education, max_education, boolean_filter.sum()))
            else:
                min_education = education_scale[i - 1]
                max_education = education_scale[i]
                boolean_filter = (min_education <= var) & (var < max_education)
                print('[{}, {})    {}'.format(min_education, max_education, boolean_filter.sum()))
            test_set = self.test_set[boolean_filter]
            test_set.index = range(test_set.shape[0])
            for model_name, scores in self.scores_dict.items():
                start_time = get_timestamp()
                eval_result = cp.compute(test_set, scores[:, boolean_filter])
                save_filename = 'education_{}_{}_{}.csv'.format(min_education, max_education, model_name)
                eval_result.to_csv(os.path.join(self.result_dir_path, save_filename), index=False)
                print('education_{}_{}_{}: {:.0f}s'.format(
                    min_education, max_education, model_name, get_timestamp() - start_time))

    def grouping_by_mmse(self):
        """
        理想分组
        [0, 11)
        [11, 21)
        [21, 26)
        [26, 30)
        [30, 30]
        """
        print('{:=^50}'.format('mmse分组'))
        print('  区间    样本量')
        var = self.test_set['mmse'].values
        mmse_scale = [0, 11, 21, 26, 30]
        for i in range(len(mmse_scale)):
            if i == len(mmse_scale) - 1:
                min_mmse = 30
                max_mmse = 30
                boolean_filter = var == 30
                print('[30, 30]    {}'.format(boolean_filter.sum()))
            else:
                min_mmse = mmse_scale[i]
                max_mmse = mmse_scale[i + 1]
                boolean_filter = (min_mmse <= var) & (var < max_mmse)
                print('[{}, {})    {}'.format(min_mmse, max_mmse, boolean_filter.sum()))

            # region 适应数据集的代码
            if boolean_filter.sum() == 0:
                continue
            if i == 1:
                nc = self.test_set['NC'] == 1
                nc_index = next(filter(lambda j: nc[j], range(nc.shape[0])))
                boolean_filter[nc_index] = True
            # endregion

            test_set = self.test_set[boolean_filter]
            test_set.index = range(test_set.shape[0])
            for model_name, scores in self.scores_dict.items():
                start_time = get_timestamp()
                eval_result = cp.compute(test_set, scores[:, boolean_filter])
                save_filename = 'mmse_{}_{}_{}.csv'.format(min_mmse, max_mmse, model_name)
                eval_result.to_csv(os.path.join(self.result_dir_path, save_filename), index=False)
                print('mmse_{}_{}_{}: {:.0f}s'.format(min_mmse, max_mmse, model_name, get_timestamp() - start_time))

    def seq_info(self, var_name, var_scale):
        var = self.test_set[var_name].values
        if var_scale[0] == 'min':
            var_scale[0] = var.min()
        if var_scale[-1] == 'max':
            var_scale[-1] = var.max()
        table = np.zeros((len(var_scale) - 1, 3), dtype='object')
        for i in range(len(var_scale) - 1):
            start = var_scale[i]
            end = var_scale[i + 1]
            table[i, 0] = var_name
            if i == len(var_scale) - 2:
                table[i, 1] = '[{}, {})'.format(start, end)
                table[i, 2] = sum((start <= var) & (var < end))
            else:
                table[i, 1] = '[{}, {}]'.format(start, end)
                table[i, 2] = sum((start <= var) & (var <= end))
        return table

    def discrete_info(self, var_name):
        var = self.test_set[var_name].values
        classes = sorted([v for v in set(var) if v not in [-4, -1]])
        table = np.zeros((len(classes), 3), dtype='object')
        for i, v in enumerate(classes):
            table[i, 0] = var_name
            table[i, 1] = v
            table[i, 2] = sum(var == v)
        return table

    def get_grouping_info(self, save_path):
        """
        将分组信息导出到excel表
        :param save_path: str. excel文件路径，若已存在，新建一个名字为“grouping_info”的sheet表存放分组信息
        :return: None
        """
        # 计算得到数据
        age = self.seq_info('age', [0, 65, 75, 85, 'max'])
        education = self.seq_info('education', [0, 12, 15, 'max'])
        family_history = self.discrete_info('family_history')
        gender = self.discrete_info('gender')
        genetics = self.discrete_info('genetics')
        mmse = self.seq_info('mmse', [0, 11, 21, 26, 30])
        data = np.vstack((age, education, family_history, gender, genetics, mmse))

        # 打开Excel表并创建grouping_info工作簿
        if os.path.exists(save_path):
            wb = openpyxl.load_workbook(save_path)
        else:
            wb = Workbook()
        if 'grouping_info' in wb.sheetnames:
            wb.remove(wb['grouping_info'])

        # 将数据写入Excel表
        info = SheetExtension(wb.create_sheet('grouping_info'))
        info[0, 0] = 'risk factor'
        info[0, 1] = 'value'
        info[0, 2] = 'count'
        for r in range(data.shape[0]):
            for c in range(data.shape[1]):
                info[r + 1, c] = data[r, c]

        # 合并单元格
        start_row = 1
        for r in range(2, data.shape[0] + 1):
            if info[start_row, 0].value != info[r, 0].value:
                info.sheet.merge_cells('A{}:A{}'.format(start_row + 1, r))
                start_row = r
        if data.shape[0] > start_row:
            info.sheet.merge_cells('A{}:A{}'.format(start_row + 1, data.shape[0] + 1))

        # 设置字体、字号、对齐方式
        for r in range(data.shape[0] + 1):
            for c in range(data.shape[1]):
                cell = info[r, c]
                if r == 0:
                    cell.font = Font(name='Consolas', size=12, bold=True)
                else:
                    cell.font = Font(name='Consolas', size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # 设置行高、列宽
        for r in range(data.shape[0] + 1):
            info.sheet.row_dimensions[r + 1].height = 15.75
        info.sheet.column_dimensions[get_column_letter(1)].width = 17.63
        info.sheet.column_dimensions[get_column_letter(2)].width = 10.13
        info.sheet.column_dimensions[get_column_letter(3)].width = 6.75

        # 设置边框
        black_side = Side(style='thin', color='000000')
        for r in range(data.shape[0] + 1):
            for c in range(data.shape[1]):
                info.set_border(r, c, black_side, 'top right bottom left')

        # 保存Excel文件并关闭
        if not os.path.exists(save_path):
            wb.remove(wb['Sheet'])
        wb.save(save_path)
        wb.close()


class SheetExtension(object):
    max_cols = 18277

    def __init__(self, sheet):
        self.sheet = sheet

    @staticmethod
    def num_to_chr(num: int) -> str:
        """
        将一个数字（从0到18277）转换成Excel表格的列字母编码（类似于26进制）
        :param num: int. 列下标，范围为[0, 17575]
        :return: str. 列字母编码
        """
        if num == 0:
            return 'A'
        if num < 0 or num > SheetExtension.max_cols:
            raise ValueError('数值{}超出最大范围：[0, {}]'.format(num, SheetExtension.max_cols))
        first_iter = True
        bits = []
        while num != 0:
            if first_iter:
                bits.append(num % 26)
                num = num // 26
                first_iter = False
            else:
                bits.append((num - 1) % 26)
                num = (num - 1) // 26
        chrs = []
        for bit in reversed(bits):
            chrs.append(chr(65 + bit))
        return ''.join(chrs)

    @staticmethod
    def chr_to_num(char: str) -> int:
        """
        将Excel表格的列字母编码（类似于26进制）转换成数字（从0到18277）
        :param char: str. 列字母编码
        :return: int.
        """
        num = 0
        for i, c in enumerate(reversed(char)):
            if i == 0:
                num = num + ord(c) - 65
            else:
                num = num + pow(26, i) * (ord(c) - 64)
        return num

    def set_border(self, r: int, c: int, side: Side, direction: str):
        """
        为单元格添加黑色边框
        :param r: int. 单元格所在行，从0开始计数
        :param c: int. 单元格所在列，从0开始计数
        :param side: openpyxl.styles.Side. 单元格边框样式对象
        :param direction: str. 例子：'left bottom right top'
        :return: None
        """
        self[r, c].border = Border(**{k: side for k in direction.split(' ')})

    def __getitem__(self, t: tuple):
        # 将索引变为slice对象
        t = list(t)
        for i, v in enumerate(t):
            if isinstance(v, int):
                t[i] = slice(v, v + 1, 1)
            elif isinstance(v, slice):
                new_v = [v.start, v.stop, v.step]
                if new_v[1] is None:
                    raise ValueError('对Excel工作簿不能取所有行或列')
                if new_v[0] is None:
                    new_v[0] = 0
                if new_v[2] is None:
                    new_v[2] = 1
                if new_v[0] < 0 or new_v[1] < 0 or new_v[2] < 0:
                    raise ValueError('索引必须大于0')
                t[i] = slice(*new_v)
            else:
                raise ValueError('不支持的参数类型')

        # 单个值切片
        rows = list(range(t[0].start, t[0].stop, t[0].step))
        cols = list(range(t[1].start, t[1].stop, t[1].step))
        if len(rows) == 1 and len(cols) == 1:
            return self.sheet.cell(row=t[0].start + 1, column=t[1].start + 1)

        # 单行或单列切片
        if len(rows) == 1 or len(cols) == 1:
            if len(rows) == 1:
                return [self.sheet.cell(row=rows[0] + 1, column=c + 1) for c in cols]
            return [self.sheet.cell(row=r + 1, column=cols[0] + 1) for r in rows]

        # 多行多列切片
        result = []
        for r in rows:
            result_row = []
            for c in cols:
                result_row.append(self.sheet.cell(row=r + 1, column=c + 1))
        return result

    def __setitem__(self, key: tuple, value: any) -> None:
        r, c = key
        real_key = '{}{}'.format(get_column_letter(c + 1), r + 1)
        self.sheet[real_key] = value


def group_compute():
    """
    对非分类型变量分组的参考文献
    age：2019 Alzheimer's Disease Facts and Figures
    education：Kukull WA, Higdon R, Bowen JD, McCormick WC, Teri L, Schellenberg GD, et al. Dementia and Alzheimer
               disease incidence: A prospective cohort study. Arch Neurol 2002;59(11):1737-46.
    MMSE：Mapping Scores Onto Stages Mini-Mental State Examination and Clinical Dementia Rating
    """
    # region 旧代码
    """
    # 确定路径
    test_set_path = os.path.join(root_path, 'lookupcsv/CrossValid/no_cross/risk_factor_test.csv')
    scores_path = {
        'MRI': os.path.join(root_path, 'model_eval/eval_result/mri/scores.npy'),
        'nonImg': os.path.join(root_path, 'model_eval/eval_result/nonImg/scores.npy'),
        'Fusion': os.path.join(root_path, 'model_eval/eval_result/Fusion/scores.npy')
    }
    result_dir_path = os.path.join(root_path, 'model_eval/eval_result/risk_factor_grouping')

    # 分组统计
    eval_obj = RiskFactorEval(test_set_path, scores_path, result_dir_path)
    # eval_obj.grouping_by_age()
    # eval_obj.grouping_by_gender()
    # eval_obj.grouping_by_education()
    # eval_obj.grouping_by_genetics()
    # eval_obj.grouping_by_family_history()
    # eval_obj.grouping_by_mmse()
    # eval_obj.get_grouping_info(os.path.join(root_path, 'model_eval/eval_result/risk_factor_show.xlsx'))
    """
    # endregion

    # 必需的路径
    performance_path = os.path.join(root_path, 'model_eval/eval_result/Fusion/result.csv')
    testset_path = os.path.join(root_path, 'lookupcsv/CrossValid/no_cross/risk_factor_test.csv')
    scores_path = os.path.join(root_path, 'model_eval/eval_result/Fusion/scores.npy')

    # 找出performance最优的模型，找出benefit最优的模型
    performance = pd.read_csv(performance_path)
    testset = pd.read_csv(testset_path)
    scores = np.load(scores_path)
    op_id, ob_id = model_selection(performance_path)
    op_pred = np.reshape(scores[op_id, :], (1, scores.shape[1]))
    ob_pred = np.reshape(scores[ob_id, :], (1, scores.shape[1]))

    def discrete(var_name, categorys):
        var = testset[var_name].values
        for i, v in enumerate(categorys):
            filter_boolean = var == v
            key = str(v)
            yield filter_boolean, key
    
    def continuous(var_name, scales):
        result[var_name] = OrderedDict()
        var = testset[var_name].values
        effective_boolean = var != -4
        for i in range(len(scales) + 1):
            if i == 0:
                start = var[effective_boolean].min()
                end = scales[i]
                filter_boolean = effective_boolean & (var < end)
                key = '[{}, {})'.format(start, end)
            elif i == len(scales):
                start = scales[-1]
                end = var[effective_boolean].max()
                filter_boolean = effective_boolean & (start <= var)
                key = '[{}, {}]'.format(start, end)
            else:
                start = scales[i - 1]
                end = scales[i]
                filter_boolean = effective_boolean & (start <= var) & (var < end)
                key = '[{}, {})'.format(start, end)
            yield filter_boolean, key

    # 分组计算：gender、education、age、family_history、genetics、mmse
    compute_config = OrderedDict({
        'gender': {
            'method': 'discrete',
            'categorys': (0, 1)
        },
        'education': {
            'method': 'continuous',
            'scales': (12, 15)
        },
        'age': {
            'method': 'continuous',
            'scales': (65, 75, 85)
        },
        'family_history': {
            'method': 'discrete',
            'categorys': (0, 1)
        },
        'genetics': {
            'method': 'discrete',
            'categorys': (0, 1, 2)
        },
        'mmse': {
            'method': 'continuous',
            'scales': (26, )
        }
    })
    result = OrderedDict()
    for var_name, config in compute_config.items():
        result[var_name] = OrderedDict()
        if config['method'] == 'discrete':
            method = discrete
            params = (var_name, config['categorys'])
        else:
            method = continuous
            params = (var_name, config['scales'])
        for filter_boolean, key in method(*params):
            filter_testset = testset[filter_boolean]

            # region 防止计算sensitivity、fpr时分母为0
            """
            出现计算sensitivity或计算fpr时分母为0的原因：
            1. 因为sensitivity = TP / (TP + FN)，分母是测试集实际标签的正例个数
               所以sensitivity分母为0是由于此次分组的测试集没有正例导致的
            2. 因为fpr = FP / (FP + TN)，分母是测试集实际标签的反例个数
               所以fpr分母为0是由于此次分组的测试集没有反例导致的
            解决方法：在对危险因素进行分组后，还要把三分类变成三个二分类，才计算指标。
                      所以"没有正例"、"没有反例"的含义在这次的场景中就是：
                      分组后的数据集标签类别数不是三种（只有两种或一种）。
                      所以当出现这种情况时，只需将缺失的标签的一个样本放入这一组数据集即可。
                      这种解决方法的缺点是分组合理性会有一点偏差，且会让指标的值变得极偏向于0或1
            """
            """
            diff = {0, 1, 2} - set(filter_testset['COG'])
            if diff:
                for v in diff:
                    arr = (testset['COG'] == v).values
                    add_index = next(filter(lambda i: arr[i], range(arr.shape[0])))
                    filter_boolean[add_index] = True
                filter_testset = testset[filter_boolean]
            """
            # endregion

            op_eval_result = cp.compute(filter_testset, op_pred[:, filter_boolean])
            ob_eval_result = cp.compute(filter_testset, ob_pred[:, filter_boolean])
            result[var_name][key] = {
                'optimal performance': op_eval_result.to_numpy().tolist()[0],
                'optimal benefit': ob_eval_result.to_numpy().tolist()[0]
            }
    return result


def write_excel(ws: SheetExtension, start_row: int, title: str, data: dict):
    # 填入值
    cols = len(data) * 2 + 1
    ws[start_row, 0] = title
    ws[start_row + 3, 0] = 'max_performance'
    ws[start_row + 4, 0] = 'max_benefit'
    for i, (var_name, values) in enumerate(data.items()):
        ws[start_row + 1, 2 * i + 1] = var_name
        ws[start_row + 2, 2 * i + 1] = 'performance'
        ws[start_row + 2, 2 * i + 2] = 'benefit'
        ws[start_row + 3, 2 * i + 1] = values[0]
        ws[start_row + 3, 2 * i + 2] = values[1]
        ws[start_row + 4, 2 * i + 1] = values[2]
        ws[start_row + 4, 2 * i + 2] = values[3]

    # 合并单元格
    ws.sheet.merge_cells('A{}:S{}'.format(start_row + 1, start_row + 1))
    for i in range(len(data)):
        ws.sheet.merge_cells('{}{}:{}{}'.format(
            get_column_letter(2 * i + 2), start_row + 2,
            get_column_letter(2 * i + 3), start_row + 2
        ))

    # 设置字体、字号、对齐方式
    font = Font(name='Consolas', size=12)
    for r in range(start_row, start_row + 5):
        for c in range(0, cols):
            cell = ws[r, c]
            cell.font = font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # 设置行高
    for r in range(start_row, start_row + 5):
        ws.sheet.row_dimensions[r + 1].height = 25.2

    # 设置列宽
    for c in range(cols):
        obj = ws.sheet.column_dimensions[get_column_letter(c + 1)]
        if c == 0:
            obj.width = 20.11
        elif c % 2 != 0:
            obj.width = 14.67
        else:
            obj.width = 13.33

    # 设置边框
    black_side = Side(style='thin', color='000000')
    for r in range(start_row, start_row + 5):
        if r == start_row:
            ws.set_border(r, 0, black_side, 'top left bottom')
        else:
            ws.set_border(r, 0, black_side, 'left bottom right')
    for c in range(1, cols):
        if c == cols - 1:
            ws.set_border(start_row, c, black_side, 'top bottom right')
        else:
            ws.set_border(start_row, c, black_side, 'top bottom')
        if c % 2 != 0:
            ws.set_border(start_row + 4, c, black_side, 'bottom')
            continue
        for r in range(start_row + 1, start_row + 4):
            ws.set_border(r, c, black_side, 'right')
        ws.set_border(start_row + 4, c, black_side, 'bottom right')


def transform_table():
    # 确定路径
    var_path = os.path.join(root_path, 'model_eval/eval_result/risk_factor_grouping')
    save_path = os.path.join(root_path, 'model_eval/eval_result/risk_factor_show.xlsx')
    
    # 整理分组计算结果
    group_result = {}
    for filename in os.listdir(var_path):
        prefix, model_name = filename[:-4].rsplit('_', 1)
        var_name, group_value = prefix.split('_', 1)
        full_path = os.path.join(var_path, filename)
        if var_name not in group_result:
            group_result[var_name] = {group_value: {model_name: full_path}}
        elif group_value not in group_result[var_name]:
            group_result[var_name][group_value] = {model_name: full_path}
        else:
            group_result[var_name][group_value][model_name] = full_path

    # 将数据写入Excel表并调整样式
    wb = Workbook()
    for var_name, d in group_result.items():
        ws = SheetExtension(wb.create_sheet(var_name))
        for i, (group_value, path_dict) in enumerate(d.items()):
            mri = pd.read_csv(path_dict['MRI'])
            nonimg = pd.read_csv(path_dict['nonImg'])
            fusion = pd.read_csv(path_dict['Fusion'])
            data = {}
            for j, performance_name in enumerate(mri.columns):
                if performance_name == 'benefit':
                    continue
                var = np.vstack((
                    mri[[performance_name, 'benefit']].to_numpy(),
                    nonimg[[performance_name, 'benefit']].to_numpy(),
                    fusion[[performance_name, 'benefit']].to_numpy()
                ))
                var = np.around(var, 4)
                max_var = var[var[:, 0].argmax(), :]
                max_benefit = var[var[:, 1].argmax(), :]
                data[performance_name] = [max_var[0], max_var[1], max_benefit[0], max_benefit[1]]
            write_excel(
                ws=ws,
                start_row=i * 8,
                title='{}_{}'.format(var_name, group_value),
                data=data
            )

    # 保存并关闭Excel文件
    wb.remove(wb['Sheet'])
    wb.save(save_path)
    wb.close()


def risk_factor_statistics(path):
    rf = pd.read_csv(path)
    cog = rf['COG'].values
    mean_std_ls = ['age', 'education', 'mmse']
    percent_ls = {'gender': 1, 'genetics': 0, 'family_history': 0}

    # 统计均值、标准差
    for var_name in mean_std_ls:
        print(var_name)
        var = rf[var_name].values
        for category_value, category_name in enumerate(category_list):
            group = var[cog == category_value]
            print(category_name, '{:.2f} ± {:.2f}'.format(group.mean(), group.std()))
        print()

    # 统计百分比
    for var_name, positive_value in percent_ls.items():
        print(var_name)
        var = rf[var_name].values
        for category_value, category_name in enumerate(category_list):
            group = var[cog == category_value]
            count = sum(group == positive_value)
            percent = count / group.shape[0]
            print(category_name, '{}({:.2%})'.format(count, percent))
        print()


def get_risk_factor_eval():
    # 确定路径
    var_path = os.path.join(root_path, 'model_eval/eval_result/risk_factor_grouping')

    # 整理分组计算结果
    group_result = {}
    for filename in os.listdir(var_path):
        prefix, model_name = filename[:-4].rsplit('_', 1)
        var_name, group_value = prefix.split('_', 1)
        full_path = os.path.join(var_path, filename)
        if var_name not in group_result:
            group_result[var_name] = {group_value: {model_name: full_path}}
        elif group_value not in group_result[var_name]:
            group_result[var_name][group_value] = {model_name: full_path}
        else:
            group_result[var_name][group_value][model_name] = full_path

    # 计算
    var_map = {
        'auc_nc': 3, 'auc_mci': 4, 'auc_de': 5, 'ap_nc': 6, 'ap_mci': 7, 'ap_de': 8,
        'sensitivity': 0, 'specificity': 1, 'accuracy': 2
    }
    result = {}
    for rf, rf_dict in group_result.items():
        result[rf] = {}
        for group_value, group_dict in rf_dict.items():
            result[rf][group_value] = {}
            group_data = np.vstack((
                pd.read_csv(group_dict['MRI']).to_numpy(),
                pd.read_csv(group_dict['nonImg']).to_numpy(),
                pd.read_csv(group_dict['Fusion']).to_numpy()
            ))
            group_data = np.around(group_data, 4)
            max_benefit_row = group_data[group_data[:, 9].argmax(), :]
            for var_name, i in var_map.items():
                result[rf][group_value][var_name] = [
                    list(group_data[group_data[:, i].argmax(), [i, 9]]),
                    list(max_benefit_row[[i, 9]])
                ]
    return result


def generate_overleaf_table(data):
    table_template = r"""\begin{table}[ht]
    \centering
    \large
    \begin{tabular}{ | c c c c c c | }
        \hline
        \multicolumn{6}{|c|}{\textbf{%title%}} \\
%content%
        \hline
    \end{tabular}
    \caption{%caption%}
    \label{table:%label%}
\end{table}"""
    content_template = r"""        \hline
        \multicolumn{6}{|l|}{\textbf{%group_value%}}\\
        $AUC_{NC}$ & $AUC_{MCI}$ & $AUC_{DE}$ & $AP_{NC}$ & $AP_{MCI}$ & $AP_{DE}$ \\
        %op_auc_nc% & %op_auc_mci% & %op_auc_de% & %op_ap_nc% & %op_ap_mci% & %op_ap_de% \\
        $Sensitivity$ & $Specificity$ & $Accuracy$ & $Benefit$ & & \\
        %op_sensitivity% & %op_specificity% & %op_accuracy% & %op_benefit% & & \\"""
    var_map = OrderedDict({
        'auc_nc': 3, 'auc_mci': 4, 'auc_de': 5,
        'ap_nc': 6, 'ap_mci': 7, 'ap_de': 8,
        'sensitivity': 0, 'specificity': 1, 'accuracy': 2,
        'benefit': 9
    })
    rf_content_map = {
        'gender': {
            'title': 'Gender',
            'caption': 'Eval by gender group',
            'label': 'eval_by_gender_group',
            'group_value': ('female', 'male')
        },
        'education': {
            'title': 'Education in years',
            'caption': 'Eval by education group',
            'label': 'eval_by_education_group',
            'group_value': ('years in [6, 12)', 'years in [12, 15)', 'years in [15, 20]')
        },
        'age': {
            'title': 'Age',
            'caption': 'Eval by age group',
            'label': 'eval_by_age_group',
            'group_value': ('age in [56, 65)', 'age in [65, 75)', 'age in [75, 85)', 'age in [85, 94]')
        },
        'family_history': {
            'title': 'Family history',
            'caption': 'Eval by family history group',
            'label': 'eval_by_family_history_group',
            'group_value': ('not', 'have')
        },
        'genetics': {
            'title': 'APOE4',
            'caption': 'Eval by APOE4 group',
            'label': 'eval_by_apoe4_group',
            'group_value': ('APOE4 = 0', 'APOE4 = 1', 'APOE4 = 2')
        },
        'mmse': {
            'title': 'MMSE',
            'caption': 'Eval by MMSE group',
            'label': 'eval_by_mmse_group',
            'group_value': ('MMSE in [11, 26)', 'MMSE in [26, 30]')
        }
    }
    for rf, rf_dict in data.items():
        rf_config = rf_content_map[rf]
        rf_content = table_template
        rf_content = rf_content.replace('%title%', rf_config['title'])
        rf_content = rf_content.replace('%caption%', rf_config['caption'])
        rf_content = rf_content.replace('%label%', rf_config['label'])
        group_content_list = []
        for group_index, (group_value, group_dict) in enumerate(rf_dict.items()):
            group_content = content_template
            group_content = group_content.replace('%group_value%', rf_config['group_value'][group_index])
            op = group_dict['optimal performance']
            ob = group_dict['optimal benefit']
            for var_name, var_index in var_map.items():
                group_content = group_content.replace('%op_{}%'.format(var_name), '{:.4f}'.format(op[var_index]))
                group_content = group_content.replace('%ob_{}%'.format(var_name), '{:.4f}'.format(ob[var_index]))
            group_content_list.append(group_content)
        rf_content = rf_content.replace('%content%', '\n\n'.join(group_content_list))
        print(rf_content)
        print()


def model_selection(performance_path: str) -> tuple:
    """
    选出Fusion的多个保存点中，最优performance和最优benefit模型
    评价最优performance的标准：(auc_nc + auc_mci + auc_de) / 3最高
    评价最优benefit的标准：benefit最高
    :param performance_path: str. 已计算的指标csv路径
    :returns: 2-tuple. (最优performance模型的文件名, 最优benefit模型的文件名)
    """
    performance = pd.read_csv(performance_path)
    op = performance.loc[0, ['auc_nc', 'auc_mci', 'auc_de']].sum() / 3  # 最优performance
    ob = performance.loc[0, 'benefit']  # 最优benefit
    op_id = 0  # 最优performance模型编号
    ob_id = 0  # 最优benefit模型编号
    for i in range(1, performance.shape[0]):
        current_performance = performance.loc[i, ['auc_nc', 'auc_mci', 'auc_de']].sum() / 3
        current_benefit = performance.loc[i, 'benefit']
        if current_performance > op:
            op = current_performance
            op_id = i
        if current_benefit > ob:
            ob = current_benefit
            ob_id = i
    return op_id, ob_id


if __name__ == '__main__':
    # preprocess(
    #     test_source_path=os.path.join(root_path, 'lookupcsv/CrossValid/no_cross/test_source.csv'),
    #     risk_factor_path=os.path.join(root_path, 'model_eval/data'),
    #     result_save_path=os.path.join(root_path, 'lookupcsv/CrossValid/no_cross/risk_factor_test.csv')
    # )

    # result = get_risk_factor_eval()
    result = group_compute()
    generate_overleaf_table(result)

    """
    ======================age分组=======================
      区间    样本量
    [56, 65)    107
    [65, 75)    481
    [75, 85)    641
    [85, 94]    153
    =====================gender分组=====================
    gender = 0	样本量 = 1345
    gender = 1	样本量 = 37
    ======================age分组=======================
      区间    样本量
    [6, 12)    54
    [12, 15)    399
    [15, 20]    922
    ====================genetics分组====================
    apoe4 = 0	样本量 = 800
    apoe4 = 1	样本量 = 454
    apoe4 = 2	样本量 = 121
    =================family_history分组=================
    family_history = 0	样本量 = 496
    family_history = 1	样本量 = 549
    ======================mmse分组======================
      区间    样本量
    [0, 11)    0
    [11, 21)    47
    [21, 26)    257
    [26, 30)    745
    [30, 30]    332
    """
